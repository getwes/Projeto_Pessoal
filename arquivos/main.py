from openpyxl import load_workbook
import openpyxl


planilha_perifericos = load_workbook('./arquivos/perifericos.xlsx')
pagina_perifericos = planilha_perifericos['perifericos']

for item_perifierico in pagina_perifericos.iter_rows(min_row=2,values_only=True):
    print(item_perifierico)


planilha_produtos = load_workbook('./arquivos/produtos.xlsx')
pagina_produtos = planilha_produtos['produtos']

for item_produto in pagina_produtos.iter_rows(min_row=2,values_only=True):
    print(item_produto)
   

